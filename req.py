# -*- coding: utf-8 -*-
"""
SmartRecruiters — Screening Q/A scraper with NER location extraction
"""

# ─── imports ────────────────────────────────────────────────────────────────
import argparse
import asyncio
import sys
import email.mime.application
import email.mime.multipart
import email.mime.text
import io
import os
import random
import re
import gc
import smtplib
import time
import unicodedata
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple
from collections import defaultdict, OrderedDict
from urllib.parse import urlparse
import concurrent.futures
import hashlib
from pathlib import Path

from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeout
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import fitz  # PyMuPDF
import geonamescache

try:
    import pycountry
except Exception:
    pycountry = None

try:
    import phonenumbers as pn
    from phonenumbers import geocoder as pn_geo
except Exception:
    pn = None
    pn_geo = None

# ─── OCR imports ────────────────────────────────────────────────────────────
try:
    from paddleocr import PaddleOCR
    from PIL import Image
    import io
    PADDLEOCR_AVAILABLE = True
    print("✅ PaddleOCR available")
except ImportError:
    PADDLEOCR_AVAILABLE = False
    print("⚠️  PaddleOCR not installed. Install with: pip install paddleocr paddlepaddle --break-system-packages")

# ─── CONFIG ─────────────────────────────────────────────────────────────────
WORKERS = 5
HUMAN_DELAY_MIN, HUMAN_DELAY_MAX = 0.18, 0.45
PAGE_LOAD_TIMEOUT = 30000
PDF_PROCESSING_TIMEOUT = 30  # Increased for slow PDF downloads and OCR processing
PDF_MAX_PAGES = 8
NAV_MIN_GAP_MS = 250
SWEEP_EVERY = 10
RETRY_LIMIT = 2
SCREENING_WAIT_AFTER_CLICK_MS = 2500

CDP_URL = "http://127.0.0.1:9222"
BASE_URL = "https://www.smartrecruiters.com"
VIEWPORT = {"width": 1920, "height": 1080}
OUTFILE = "smartrecruiters_screening_all_QA.xlsx"

# Email config (set EMAIL_USER, EMAIL_PASSWORD, EMAIL_PROVIDER in env)
# EMAIL_PROVIDER: "gmail" or "outlook" — use "gmail" (Outlook Basic Auth often disabled)
OUTLOOK_SMTP = "smtp.office365.com"
OUTLOOK_PORT = 587
GMAIL_SMTP = "smtp.gmail.com"
GMAIL_PORT = 587

# Location scoring config
SCORE_HIGH = 10
SCORE_MED = 7
SCORE_MIN = 8
STRICT_LOC_MODE = True
REQUIRE_HEADER = True
REQUIRE_COUNTRY_MATCH = True
LOCATION_DEBUG = True

# NER Configuration
USE_NER = True
NER_MIN_CONFIDENCE = 0.3  # Lowered from 0.4 to catch more locations
NER_DEBUG = True

# OCR Configuration
USE_OCR = PADDLEOCR_AVAILABLE  # Auto-enable if available
OCR_MAX_PAGES = 2  # Only OCR first 2 pages (contact info usually there)
OCR_MIN_TEXT_LENGTH = 50  # Only OCR if text extraction got <50 chars
OCR_SKIP_IF_GOOD_DOM = True  # Skip OCR if DOM has detailed location (city + state)
OCR_CACHE_DIR = Path(".ocr_cache")  # Cache OCR results to avoid reprocessing
OCR_DEBUG = True
# When True: always run OCR on first page of every PDF (to verify OCR works; adds ~2–5s per profile)
OCR_DEBUG_ALWAYS_FIRST_PAGE = False

# Initialize OCR cache directory
if USE_OCR:
    OCR_CACHE_DIR.mkdir(exist_ok=True)

# Initialize PaddleOCR (lazy loading)
_OCR_ENGINE = None

def get_ocr_engine():
    """Lazy load OCR engine on first use"""
    global _OCR_ENGINE
    if _OCR_ENGINE is None and PADDLEOCR_AVAILABLE:
        print("📦 Loading PaddleOCR engine...")
        _OCR_ENGINE = PaddleOCR(
            use_angle_cls=True,  # Detect text orientation
            lang='en',           # English language
            show_log=False,      # Suppress verbose logs
            use_gpu=False        # CPU mode (set True if you have CUDA)
        )
        print("✅ PaddleOCR ready")
    return _OCR_ENGINE

# Load NER model
_NER_EXTRACTOR = None
try:
    # using improved NER with Indian location awareness
    from ner_location_improved import ImprovedLocationNER as LocationNER, format_location
    print("📦 Loading improved NER model...")
    _NER_EXTRACTOR = LocationNER()
    print("✅ Improved NER ready (with Indian location support)")
except ImportError:
    # Fallback to original NER if improved version not available
    try:
        from ner_location import ImprovedLocationNER as LocationNER, format_location
        print("📦 Loading NER model...")
        _NER_EXTRACTOR = LocationNER()
        print("✅ NER ready (standard version)")
    except Exception as e:
        USE_NER = False
        print(f"⚠️ NER not available: {e}")

# ─── USER INPUT: keywords to flag (Ctrl+F style) ─────────────────────────────
# Default fallback when no --keywords, --keywords-file, or prompt
DEFAULT_KEYWORDS: List[str] = [
    "Pytorch", "pytorch", "PyTorch", "Py-Torch",
    "Tensorflow", "tensorflow", "keras", "Keras",
    "Azure", "cloud", "aws", "azure", "PHD",
    "Chennai", "chennai", "Pune", "pune",
    "Research", "research", "Machine Learning", "machine learning",
    "Deep Learning", "deep learning",
    "NLP", "nlp", "natural language processing",
    "r&D", "R&D", "research & development", "Research & Development",
]

# Abbreviation → full forms / synonyms (resilience: catch variations like Ctrl+F would miss)
_RD_EXPANSIONS = ["research and development", "research & development", "r and d", "r&d"]
KEYWORD_EXPANSIONS: Dict[str, List[str]] = {
    "r&d": _RD_EXPANSIONS,
    "r and d": _RD_EXPANSIONS,
    "r & d": _RD_EXPANSIONS,
    "ml": ["machine learning", "machine-learning"],
    "nlp": ["natural language processing", "natural-language processing"],
    "ai": ["artificial intelligence", "artificial-intelligence"],
    "dl": ["deep learning", "deep-learning"],
    "cv": ["computer vision"],
    "phd": ["ph.d", "ph.d.", "doctorate", "doctoral"],
    "ms": ["m.s", "m.s.", "master's", "masters", "msc", "m.sc"],
    "bsc": ["b.s", "b.s.", "bachelor's", "bachelors", "b.sc"],
    "iso 45001": ["iso45001", "iso-45001", "ohsms", "occupational health and safety"],
    "iso 9001": ["iso9001", "iso-9001", "quality management"],
    "nebsh": ["nebsh igc", "international general certificate"],
    "ctf": ["capture the flag", "capture-the-flag"],
    "aws": ["amazon web services"],
    "gcp": ["google cloud platform", "google cloud"],
    "api": ["application programming interface", "apis"],
}


def _load_custom_expansions() -> None:
    """Load extra expansions from keyword_expansions.txt if it exists. Format: abbr|full1, full2"""
    path = Path(__file__).parent / "keyword_expansions.txt"
    if not path.exists():
        return
    try:
        for line in path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "|" in line:
                abbr, _, rest = line.partition("|")
                abbr = abbr.strip().lower()
                forms = [f.strip() for f in rest.split(",") if f.strip()]
                if abbr and forms:
                    KEYWORD_EXPANSIONS.setdefault(abbr, []).extend(forms)
    except Exception:
        pass


_load_custom_expansions()


def expand_keywords(keywords: List[str]) -> List[str]:
    """
    Add common full forms / synonyms for abbreviations.
    E.g. R&D → also search "research and development".
    """
    expanded = list(keywords)
    expanded_lower = {k.lower() for k in expanded}
    for kw in keywords:
        key = (kw or "").strip().lower()
        if not key:
            continue
        if key in KEYWORD_EXPANSIONS:
            for form in KEYWORD_EXPANSIONS[key]:
                if form.lower() not in expanded_lower:
                    expanded.append(form)
                    expanded_lower.add(form.lower())
    return expanded


def canonicalize_keywords(keywords: List[str]) -> List[str]:
    """
    - Strip whitespace
    - Drop empty entries
    - Deduplicate case-insensitively
    - Keep the first-seen casing as the display form
    """
    seen_lower = set()
    cleaned: List[str] = []
    for kw in keywords:
        kw_clean = (kw or "").strip()
        if not kw_clean:
            continue
        key = kw_clean.lower()
        if key in seen_lower:
            continue
        seen_lower.add(key)
        cleaned.append(kw_clean)
    return cleaned


def parse_keywords_from_string(s: str) -> List[str]:
    """Parse keywords from string: comma, newline, or semicolon separated. Lines starting with # are ignored."""
    if not s or not s.strip():
        return []
    lines = s.splitlines()
    cleaned = []
    for line in lines:
        if line.strip().startswith("#"):
            continue
        for part in re.split(r"[,;]+", line):
            p = part.strip()
            if p:
                cleaned.append(p)
    return cleaned


def _resolve_and_expand(parsed: List[str]) -> List[str]:
    """Canonicalize, expand abbreviations, then dedupe again."""
    canon = canonicalize_keywords(parsed)
    expanded = expand_keywords(canon)
    return canonicalize_keywords(expanded)


def resolve_keywords(
    keywords_arg: Optional[str],
    keywords_file_arg: Optional[str],
) -> List[str]:
    """
    Resolve keywords from CLI, file, or interactive prompt.
    Order: --keywords > --keywords-file > interactive prompt > DEFAULT_KEYWORDS
    Abbreviations (R&D, ML, NLP, etc.) are auto-expanded to full forms for resilience.
    """
    if keywords_arg:
        parsed = parse_keywords_from_string(keywords_arg)
        if parsed:
            return _resolve_and_expand(parsed)
    env_kw = os.environ.get("KEYWORDS", "").strip()
    if env_kw:
        parsed = parse_keywords_from_string(env_kw)
        if parsed:
            return _resolve_and_expand(parsed)
    file_path = keywords_file_arg or os.environ.get("KEYWORDS_FILE")
    if file_path:
        path = Path(file_path)
        if path.exists():
            try:
                text = path.read_text(encoding="utf-8")
                parsed = parse_keywords_from_string(text)
                if parsed:
                    return _resolve_and_expand(parsed)
            except Exception as e:
                print(f"⚠️  Could not read keywords file: {e}")
    if sys.stdin.isatty():
        print("\n🔍 Keywords (Ctrl+F style): Enter terms to search in resumes.")
        print("   Comma or newline separated. Abbreviations (R&D, ML, NLP) auto-expand.")
        print("   Examples: Python, ML, AWS, R&D, Chennai. Enter alone for defaults.\n")
        lines = []
        try:
            while True:
                line = input("   Keyword(s)> ").strip()
                if not line:
                    break
                lines.append(line)
        except (EOFError, KeyboardInterrupt):
            pass
        if lines:
            parsed = parse_keywords_from_string("\n".join(lines))
            if parsed:
                return _resolve_and_expand(parsed)
    return _resolve_and_expand(DEFAULT_KEYWORDS)


def _sanitize_email_str(s: str) -> str:
    """Remove NBSP and other chars that break ASCII encoding in email headers."""
    if not s:
        return ""
    return s.replace("\xa0", " ").replace("\u00a0", " ").encode("ascii", "ignore").decode("ascii")


def _load_email_credentials() -> None:
    """Load EMAIL_* from email_credentials.env if it exists (avoids terminal copy-paste issues)."""
    creds_file = Path(__file__).parent / "email_credentials.env"
    if not creds_file.exists():
        return
    try:
        for line in creds_file.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" in line:
                key, _, value = line.partition("=")
                key = key.strip()
                value = value.strip().strip("'\"")
                if key in ("EMAIL_USER", "EMAIL_PASSWORD", "EMAIL_PROVIDER", "EMAIL_TO", "KEYWORDS_FILE") and value:
                    os.environ.setdefault(key, value)
    except Exception:
        pass


def _send_email_smtp(
    to_email: str,
    excel_bytes: bytes,
    smtp_host: str,
    smtp_port: int,
    filename: str = "smartrecruiters_report.xlsx",
    subject: str = "SmartRecruiters Prospect Report",
) -> bool:
    """Send Excel report via SMTP. Uses EMAIL_USER and EMAIL_PASSWORD from env."""
    user = _sanitize_email_str(os.environ.get("EMAIL_USER", "").strip())
    password = re.sub(r"\s+", "", os.environ.get("EMAIL_PASSWORD", ""))  # strip ALL whitespace (spaces, NBSP, etc.)
    to_email = _sanitize_email_str(to_email)
    subject = _sanitize_email_str(subject)
    filename = _sanitize_email_str(filename)
    if not user or not password:
        print("⚠️  Set EMAIL_USER and EMAIL_PASSWORD in environment to send email.")
        return False
    try:
        msg = email.mime.multipart.MIMEMultipart()
        msg["From"] = user
        msg["To"] = to_email
        msg["Subject"] = subject
        msg.attach(email.mime.text.MIMEText("Report attached.", "plain", "utf-8"))
        part = email.mime.application.MIMEApplication(excel_bytes, _subtype="xlsx")
        part.add_header("Content-Disposition", "attachment", filename=filename)
        msg.attach(part)
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(user, password)
            server.sendmail(user, to_email, msg.as_string())
        print(f"📧 Report sent to {to_email}")
        return True
    except Exception as e:
        print(f"⚠️  Email failed: {e}")
        return False


def send_email(
    to_email: str,
    excel_bytes: bytes,
    filename: str = "smartrecruiters_report.xlsx",
    subject: str = "SmartRecruiters Prospect Report",
) -> bool:
    """
    Send Excel report via email. Uses EMAIL_PROVIDER to choose SMTP:
    - "gmail": Gmail SMTP (App Password required, 2FA must be on)
    - "outlook": Outlook SMTP (Basic Auth often disabled by Microsoft)
    Credentials: env vars or email_credentials.env (avoids terminal copy-paste issues).
    """
    _load_email_credentials()
    provider = os.environ.get("EMAIL_PROVIDER", "").strip().lower()
    if not provider and os.environ.get("EMAIL_USER", "").lower().count("gmail"):
        provider = "gmail"
    if provider != "gmail":
        provider = "outlook"

    if provider == "gmail":
        return _send_email_smtp(to_email, excel_bytes, GMAIL_SMTP, GMAIL_PORT, filename, subject)
    return _send_email_smtp(to_email, excel_bytes, OUTLOOK_SMTP, OUTLOOK_PORT, filename, subject)


# ─── OPTIONAL: semantic keyword tagging (embeddings) ─────────────────────────
# Requires: pip install torch sentence-transformers numpy
# And a file named `semantic_tagger.py` in the same directory as this script.
SEMANTIC_TAGGING = True
SEMANTIC_THRESHOLD = 0.38   # 0.34–0.36 = more recall, 0.42+ = stricter
SEMANTIC_TOP_K = 6

try:
    from semantic_tagger import SemanticTagger, TagSpec

    TAG_SPECS = [
        TagSpec(
            tag="ISO 45001",
            description="Occupational health and safety management system standard; workplace safety management and compliance.",
            synonyms=("OHSMS", "occupational health and safety", "safety management system", "ISO45001"),
        ),
        TagSpec(
            tag="NIOSH",
            description="NIOSH guidance/standards; industrial hygiene, exposure limits, workplace health/safety best practices.",
            synonyms=("industrial hygiene", "exposure limits", "respiratory protection", "occupational health"),
        ),
        TagSpec(
            tag="NEBOSH",
            description="NEBOSH certification; occupational safety and health qualification.",
            synonyms=("NEBOSH IGC", "international general certificate", "safety certification"),
        ),
        TagSpec(
            tag="DOSH",
            description="Department of Occupational Safety and Health; DOSH compliance, registrations, regulations (often Malaysia).",
            synonyms=("JKKP", "occupational safety and health department", "safety regulation compliance"),
        ),
        TagSpec(
            tag="Riskonnect",
            description="Riskonnect platform; risk management, incident management, GRC tooling.",
            synonyms=("risk management system", "incident management platform", "GRC tool"),
        ),
        TagSpec(
            tag="JIRA",
            description="JIRA usage; ticketing, agile workflows, issue tracking.",
            synonyms=("Atlassian", "issue tracking", "tickets", "scrum", "kanban"),
        ),
        TagSpec(
            tag="SaaS",
            description="Software-as-a-Service; cloud-delivered applications/products.",
            synonyms=("cloud software", "subscription software", "software platform"),
        ),
    ]

    # Build once per run; reuse for all profiles
    SEM_TAGGER = SemanticTagger(TAG_SPECS)

except Exception as _e:
    SEMANTIC_TAGGING = False
    SEM_TAGGER = None
    print(f"⚠️  Semantic tagger not available (semantic_tagger.py / deps missing): {_e}")


# ─── Gazetteer (GLOBAL) ─────────────────────────────────────────────────────
_gc = geonamescache.GeonamesCache()

_COUNTRIES = {c["name"].lower(): c for c in _gc.get_countries().values()}
_COUNTRY_ALIASES = {
    "uk": "United Kingdom", "gb": "United Kingdom", "great britain": "United Kingdom",
    "uae": "United Arab Emirates", "u.a.e": "United Arab Emirates",
    "usa": "United States", "us": "United States", "u.s.": "United States", "u.s.a.": "United States",
    "czech republic": "Czechia", "russia": "Russian Federation",
    "south korea": "Korea, Republic of", "north korea": "Korea, Democratic People's Republic of",
    "viet nam": "Viet Nam"
}

def _country_lookup(name: str):
    n = re.sub(r"[^\w\s\-\.]", "", (name or "").lower()).strip()
    n = _COUNTRY_ALIASES.get(n, n)
    return _COUNTRIES.get(n)

# city index with alternatenames support
_CITIES: Dict[str, List[Dict]] = {}
for cid, c in _gc.get_cities().items():
    nm = c["name"].lower()
    _CITIES.setdefault(nm, []).append(c)
    alt = c.get("alternatenames") or []
    if not isinstance(alt, list):
        alt = str(alt).split(",")
    for a in alt:
        a = str(a).strip().lower()
        if a:
            _CITIES.setdefault(a, []).append(c)

# ─── helpers ────────────────────────────────────────────────────────────────
def _clean(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())

def strip_contacts_and_noise(s: str) -> str:
    s = re.sub(r"\S+@\S+", " ", s)
    s = re.sub(r"https?://\S+", " ", s)
    s = re.sub(r"\b(email|e-mail|phone|mobile|tel|github|linkedin|portfolio|website)\b[:\-]?", " ", s, flags=re.I)
    return _clean(s)

def _is_real_city(token: str) -> bool:
    t = token.lower()
    if t in _COUNTRIES:
        return False
    return t in _CITIES

def _city_countries(token: str) -> set:
    return {rec["countrycode"] for rec in _CITIES.get(token.lower(), [])}

def _country_name_from_iso2(iso2: str) -> str:
    for c in _COUNTRIES.values():
        if c.get("iso") == iso2:
            return c["name"]
    return ""

def _expand_subdivision(token: str, country_iso2: str) -> Optional[str]:
    if not pycountry or not country_iso2:
        return None
    name = token.strip()
    try:
        for s in pycountry.subdivisions.get(country_code=country_iso2):
            if s.name.lower() == name.lower():
                return s.name
            if s.code.endswith("-" + name.upper()):
                return s.name
    except Exception:
        pass
    return None

def _compose(city: str, region: Optional[str], country_iso2: str) -> str:
    parts = [city.title()]
    if region:
        parts.append(region)
    country_name = _country_name_from_iso2(country_iso2)
    if country_name:
        parts.append(country_name)
    return ", ".join(parts)

_STOPWORDS = {"of", "and", "or", "in", "for", "with", "to", "on", "at", "by", "from"}

_TECH_BLOCK = {t.lower() for t in """
sql git nlp cnn ml ai dl tlp rpa etl ocr api rest grpc json xml soap jira tableau powerbi power bi
excel pandas numpy pytorch tensorflow keras opencv spark hadoop kafka airflow kubernetes docker
aws gcp azure react angular vue django flask spring kotlin swift golang c c++ c# php html css js
""".split()}

def _looks_like_tech(token: str) -> bool:
    t = re.sub(r"[^\w]", "", token or "").lower()
    return t in _TECH_BLOCK or (t.isupper() and 2 <= len(t) <= 4)

# ─── Unicode-friendly résumé patterns ───────────────────────────────────────
UL = r"[^\W\d_]"

LABEL_RE = re.compile(
    rf"\b(?:location|city|address|based in|residing in|residence|locality|place)\b[:\-—]?\s*({UL}|[ ,.\-']){{2,120}}",
    re.I
)

PAIR_RE = re.compile(
    r"\b([A-Za-z][A-Za-z .'\-]{1,40})\s*,\s*([A-Za-z][A-Za-z .'\-]{1,40})(?:\s*,\s*([A-Za-z][A-Za-z .'\-]{1,40}))?\b",
    re.I | re.U
)

def normalize_location_string(raw: str) -> str:
    s = strip_contacts_and_noise(raw)
    if not s:
        return ""
    tokens = [t.strip(" .") for t in re.split(r"[,\u2013|/;-]+", s) if t.strip(" .")]
    if len(tokens) < 2:
        return ""
    country = _country_lookup(tokens[-1])
    if not country:
        return ""
    city_idx, city = -1, ""
    for i, tok in enumerate(tokens[:-1]):
        low = tok.lower()
        if _looks_like_tech(tok):
            continue
        if len(tok) < 3 or low in _STOPWORDS:
            continue
        if _is_real_city(tok):
            city_idx, city = i, tok
            break
    if not city:
        return ""
    if country["iso"] not in _city_countries(city):
        return ""
    region = None
    if city_idx + 1 < len(tokens) - 1:
        reg_name = _expand_subdivision(tokens[city_idx + 1], country["iso"])
        if reg_name:
            region = reg_name
    return _compose(city, region, country["iso"])

# ─── scoring infra ──────────────────────────────────────────────────────────
_LOCATION_LABELS = {
    "location", "based", "based in", "city", "town", "country", "address",
    "current location", "present location", "presently in", "resides in",
    "living in", "lives in", "residence", "home", "base", "place"
}
_EXPERIENCE_HEADERS = {
    "experience", "work experience", "employment", "work history",
    "professional experience", "projects"
}

def _ascii_lower(s: str) -> str:
    return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii").lower()

def _tokenize_lines(text: str) -> List[str]:
    text = text.replace("\r", "")
    return [l.strip() for l in text.split("\n") if l.strip()]

def _ngrams(tokens: List[str], n: int) -> List[str]:
    return [" ".join(tokens[i:i+n]) for i in range(0, len(tokens) - n + 1)]

def _split_words(line: str) -> List[str]:
    return re.findall(r"[a-zA-Z][a-zA-Z\.-]+|\d+", line)

def _has_label_near(line_lower: str) -> bool:
    return any(lbl in line_lower for lbl in _LOCATION_LABELS)

def _is_experience_header(line_lower: str) -> bool:
    l = re.sub(r"[^a-z ]+", " ", line_lower)
    l = re.sub(r"\s+", " ", l).strip()
    return l in _EXPERIENCE_HEADERS

@dataclass
class CityRec:
    name: str
    country_code: str
    lat: float
    lon: float
    population: int

class _GeoIndex:
    def __init__(self):
        self.country_by_name: Dict[str, Dict] = {}
        self.country_by_code: Dict[str, Dict] = {}
        self.city_by_name: Dict[str, List[CityRec]] = {}

    def build(self):
        GC = geonamescache.GeonamesCache()
        for _, c in GC.get_countries().items():
            name = _ascii_lower(c["name"])
            self.country_by_name[name] = c
            self.country_by_code[c["iso"]] = c
        for alias, canon in {
            "uk": "United Kingdom", "u.k.": "United Kingdom", "great britain": "United Kingdom",
            "usa": "United States", "u.s.a.": "United States", "us": "United States",
            "u.s.": "United States", "uae": "United Arab Emirates"
        }.items():
            self.country_by_name[alias] = self.country_by_name.get(_ascii_lower(canon), {})
        for city in GC.get_cities().values():
            nm = _ascii_lower(city["name"])
            self.city_by_name.setdefault(nm, []).append(
                CityRec(city["name"], city["countrycode"], float(city["latitude"]),
                        float(city["longitude"]), int(city.get("population") or 0))
            )

    def match_country(self, token: str) -> Optional[Dict]:
        return self.country_by_name.get(_ascii_lower(token))

    def match_city(self, token: str) -> List[CityRec]:
        return self.city_by_name.get(_ascii_lower(token), [])

_GEO = _GeoIndex()
_GEO.build()

@dataclass
class Candidate:
    city: CityRec
    line_idx: int
    line_text: str
    score: int = 0
    reasons: List[str] = None

    def bump(self, pts: int, why: str):
        self.score += pts
        if self.reasons is None:
            self.reasons = []
        self.reasons.append(why)

def _collect_candidates(lines: List[str]) -> Tuple[List[Candidate], Dict[int, str]]:
    candidates: List[Candidate] = []
    country_by_line: Dict[int, str] = {}
    for i, line in enumerate(lines):
        line_ascii = _ascii_lower(line)
        tokens = _split_words(line_ascii)
        if not tokens:
            continue
        for n in range(3, 0, -1):
            for ng in _ngrams(tokens, n):
                c = _GEO.match_country(ng)
                if c:
                    country_by_line[i] = c["iso"]
                    break
            if i in country_by_line:
                break
        seen = set()
        for n in (3, 2, 1):
            for ng in _ngrams(tokens, n):
                if ng in seen:
                    continue
                cities = _GEO.match_city(ng)
                if not cities:
                    continue
                seen.add(ng)
                for city in cities:
                    candidates.append(Candidate(city=city, line_idx=i, line_text=line))
    return candidates, country_by_line

def _phone_hints_by_line(lines: List[str]) -> Dict[int, List[Tuple[str, str]]]:
    hints: Dict[int, List[Tuple[str, str]]] = {}
    if not pn:
        return hints
    for idx, line in enumerate(lines):
        try:
            for m in pn.PhoneNumberMatcher(line, None):
                num = m.number
                if not pn.is_possible_number(num) or not pn.is_valid_number(num):
                    continue
                iso = (pn.region_code_for_number(num) or "").upper()
                if not iso:
                    continue
                loc = ""
                try:
                    loc = (pn_geo.description_for_number(num, "en") or "").strip()
                except Exception:
                    pass
                hints.setdefault(idx, []).append((iso, _ascii_lower(loc)))
        except Exception:
            continue
    return hints

def _phone_iso_votes(lines: List[str]) -> Optional[str]:
    if not pn:
        return None
    counts: Dict[str, int] = {}
    for hits in _phone_hints_by_line(lines).values():
        for iso, _ in hits:
            counts[iso] = counts.get(iso, 0) + 1
    if not counts:
        return None
    return max(counts.items(), key=lambda kv: kv[1])[0]

_CC_TLD_TO_ISO = {
    "al": "AL", "ba": "BA", "bg": "BG", "cz": "CZ", "de": "DE", "dk": "DK",
    "ee": "EE", "es": "ES", "fi": "FI", "fr": "FR", "gr": "GR", "hr": "HR",
    "hu": "HU", "ie": "IE", "it": "IT", "lt": "LT", "lv": "LV", "nl": "NL",
    "no": "NO", "pl": "PL", "pt": "PT", "ro": "RO", "rs": "RS", "se": "SE",
    "si": "SI", "sk": "SK", "tr": "TR", "ua": "UA", "uk": "GB"
}
_EMAIL_RE = re.compile(r"\b[A-Z0-9._%+\-]+@([A-Z0-9.\-]+\.[A-Z]{2,})\b", re.I)

def _email_iso_votes(lines: List[str]) -> Optional[str]:
    counts: Dict[str, int] = {}
    for line in lines:
        for m in _EMAIL_RE.finditer(line):
            tld = m.group(1).lower().rsplit(".", 1)[-1]
            iso = _CC_TLD_TO_ISO.get(tld)
            if iso:
                counts[iso] = counts.get(iso, 0) + 1
    if not counts:
        return None
    return max(counts.items(), key=lambda kv: kv[1])[0]

def _score_candidates(cands: List[Candidate], lines: List[str], country_by_line: Dict[int, str]) -> None:
    if not cands:
        return
    header_char_limit = min(700, sum(len(l) for l in lines) // 6)
    header_lines = set()
    acc = 0
    for idx, l in enumerate(lines):
        acc += len(l) + 1
        if acc <= header_char_limit:
            header_lines.add(idx)
        else:
            break
    freq: Dict[Tuple[str, str], int] = {}
    for c in cands:
        key = (c.city.name.lower(), c.city.country_code)
        freq[key] = freq.get(key, 0) + 1
    exp_start_idx = None
    for i, l in enumerate(lines):
        if _is_experience_header(_ascii_lower(l)):
            exp_start_idx = i
            break
    phone_hints = _phone_hints_by_line(lines)
    email_iso = _email_iso_votes(lines)
    for c in cands:
        line_l = _ascii_lower(c.line_text)
        if c.city.population >= 5_000_000:
            c.bump(2, "pop>5M")
        elif c.city.population >= 1_000_000:
            c.bump(1, "pop>1M")
        if c.line_idx in header_lines:
            c.bump(5, "header")
        if _has_label_near(line_l):
            c.bump(3, "label")
        iso = country_by_line.get(c.line_idx)
        if iso:
            c.bump(4 if iso == c.city.country_code else 1,
                   "country_match" if iso == c.city.country_code else "country_present")
        for ph_iso, ph_loc in phone_hints.get(c.line_idx, []):
            if ph_iso == c.city.country_code:
                c.bump(2, "phone_cc_match")
                if ph_loc and _ascii_lower(c.city.name) in ph_loc:
                    c.bump(2, "phone_area_match")
        if email_iso and email_iso == c.city.country_code:
            c.bump(1, "email_tld_match")
        key = (c.city.name.lower(), c.city.country_code)
        if freq.get(key, 0) > 1:
            c.bump(min(4, freq[key] - 1), "repeat")
        if exp_start_idx is not None and c.line_idx >= exp_start_idx:
            c.bump(-2, "in_experience")
        if re.search(rf"\b{re.escape(_ascii_lower(c.city.name))}\b\s*[,|\-]\s*[a-z ]+$", line_l):
            c.bump(2, "city_country_pattern")

def _pick_best(cands: List[Candidate]) -> Optional[Candidate]:
    if not cands:
        return None
    cands = sorted(cands, key=lambda c: (c.score, c.city.population), reverse=True)
    top = cands[0]
    if top.score < SCORE_MIN:
        return None
    if STRICT_LOC_MODE:
        rs = set(top.reasons or [])
        if REQUIRE_HEADER and "header" not in rs:
            return None
        if REQUIRE_COUNTRY_MATCH and not (("country_match" in rs) or ("city_country_pattern" in rs)):
            return None
    return top

def _normalize_country(country_iso: str) -> Tuple[str, str]:
    for _, c in _gc.get_countries().items():
        if c["iso"] == country_iso:
            return c["name"], c["iso"]
    if pycountry:
        try:
            cp = pycountry.countries.get(alpha_2=country_iso)
            if cp:
                return cp.name, cp.alpha_2
        except Exception:
            pass
    return country_iso, country_iso

def infer_location_from_text_details(text: str):
    if not text or len(text) < 20:
        return None
    lines = _tokenize_lines(text)
    cands, country_by_line = _collect_candidates(lines)
    if not cands:
        return None
    _score_candidates(cands, lines, country_by_line)
    best = _pick_best(cands)
    if not best:
        return None
    country_name, country_iso = _normalize_country(best.city.country_code)
    confidence = ("high" if best.score >= SCORE_HIGH else
                  "medium" if best.score >= SCORE_MED else "low")
    return {
        "city": best.city.name, "country": country_name, "country_code": country_iso,
        "lat": best.city.lat, "lon": best.city.lon, "confidence": confidence,
        "score": best.score, "reasons": best.reasons or []
    }

def _header_single_city_in_iso(lines: List[str], iso: str) -> Optional[str]:
    header_char_limit = min(700, sum(len(l) for l in lines) // 6)
    header_lines = []
    acc = 0
    for l in lines:
        acc += len(l) + 1
        if acc <= header_char_limit:
            header_lines.append(l)
        else:
            break
    cities = set()
    for l in header_lines:
        tokens = _split_words(_ascii_lower(l))
        for n in (3, 2, 1):
            for ng in _ngrams(tokens, n):
                for c in _GEO.match_city(ng):
                    if c.country_code == iso:
                        cities.add(c.name)
    if len(cities) == 1:
        return list(cities)[0]
    best = None
    pop = -1
    for nm in cities:
        for c in _GEO.match_city(nm):
            if c.country_code == iso and c.population > pop:
                best = nm
                pop = c.population
    return best

def infer_location_from_resume_details(text: str, dom_city_only: Optional[str] = None):
    """Regex-based location extraction"""
    head = "\n".join(_clean(text).splitlines()[:100])
    for m in LABEL_RE.finditer(head):
        cand = normalize_location_string(m.group(1))
        if cand:
            return cand, {"method": "regex", "confidence": 0.99, "score": 99, "reasons": ["dom_label"]}
    for line in head.splitlines():
        if re.search(r"\b(skills?|technolog(y|ies)|tools?|frameworks?)\b", line, re.I):
            continue
        for m in PAIR_RE.finditer(line):
            tokens = [x for x in m.groups() if x]
            if not tokens or len(tokens) < 2:
                continue
            if not _country_lookup(tokens[-1]):
                continue
            cand = normalize_location_string(", ".join(tokens))
            if cand:
                return cand, {"method": "regex", "confidence": 0.90, "score": 90, "reasons": ["pair_city_country"]}
    d = infer_location_from_text_details(text)
    if d:
        d["method"] = "regex"
        return f"{d['city']}, {d['country']}", d

    lines = _tokenize_lines(text)
    phone_iso = _phone_iso_votes(lines)
    email_iso = _email_iso_votes(lines)
    iso = phone_iso or email_iso
    if not iso:
        return "", None

    city = _header_single_city_in_iso(lines, iso)
    if not city and dom_city_only and _is_real_city(dom_city_only) and iso in _city_countries(dom_city_only):
        city = dom_city_only

    if city:
        return f"{city}, {_country_name_from_iso2(iso)}", {
            "method": "regex", "confidence": 0.70, "score": SCORE_MIN,
            "reasons": ["header_city+phone_or_email_iso" + ("" if not dom_city_only else "+dom_city")]
        }
    return "", None

# ─── HYBRID: NER + Regex ────────────────────────────────────────────────────


def _extract_contact_block(text: str, max_lines: int = 80) -> str:
    """Extract a likely 'Contact / Header' section from a resume for fast, high-precision regex parsing."""
    lines = _tokenize_lines(text)
    if not lines:
        return ""
    stop_re = re.compile(r"(?i)^(experience|work history|employment|professional experience|education|skills|summary|profile)\b")
    out = []
    for ln in lines:
        s = (ln or "").strip()
        if not s:
            continue
        if stop_re.match(s):
            break
        out.append(s)
        if len(out) >= max_lines:
            break

    # Also pull any explicit 'Address/Location/Based' lines from anywhere in the resume
    label_re = re.compile(r"(?i)\b(address|location|based in|city|country|residing|residence)\b")
    extras = []
    for ln in lines:
        s = (ln or "").strip()
        if s and label_re.search(s):
            extras.append(s)

    # de-dupe while preserving order
    seen = set()
    merged = []
    for ln in out + extras:
        key = ln.lower()
        if key in seen:
            continue
        seen.add(key)
        merged.append(ln)
    return "\n".join(merged)


def _country_from_phone(phone: str) -> str:
    """Best-effort country inference from an E.164 phone number."""
    if not phone:
        return ""
    if not pn:
        return ""
    try:
        num = pn.parse(phone, None)
        iso2 = pn.region_code_for_number(num) or ""
        iso2 = iso2.strip().upper()
        if not iso2:
            return ""
        return _country_name_from_iso2(iso2)
    except Exception:
        return ""


async def infer_location_from_resume(resume_text: str, phone: str = "", dom_city_only: str = ""):
    """
    Explicit pipeline:
      1) regex contact/address section first
      2) NER + gazetteer on full resume
      3) phone country fallback
    Returns: (location_str, meta_dict)
    """
    resume_text = resume_text or ""
    if not resume_text.strip():
        # Phone-only fallback (rare but better than empty)
        c = _country_from_phone(phone)
        if c:
            return c, {"method": "phone_country_fallback", "score": 40, "confidence": 0.40, "reasons": ["pipeline_stage=3"]}
        return "", None

    # ---- Stage 1: Contact/header regex ----
    contact = _extract_contact_block(resume_text)
    contact_plus = contact + (f"\nPhone: {phone}" if phone else "")
    loc1, meta1 = infer_location_from_resume_details(contact_plus, dom_city_only=dom_city_only)
    if loc1:
        meta1 = dict(meta1 or {})
        meta1["method"] = "resume_regex_contact"
        meta1["reasons"] = list(meta1.get("reasons") or []) + ["pipeline_stage=1"]
        return loc1, meta1

    # ---- Stage 2: NER + gazetteer on full resume ----
    full = resume_text + (f"\nPhone: {phone}" if phone else "")

    # 2a) NER extractor (spacy) if available
    if USE_NER and _NER_EXTRACTOR:
        try:
            loop = asyncio.get_running_loop()
            ner_loc = await loop.run_in_executor(None, _NER_EXTRACTOR.extract_from_resume, full)
            # expected: {'city':..., 'region':..., 'country':..., 'confidence': float, 'method': 'ner', ...}
            if isinstance(ner_loc, dict):
                cand = format_location(ner_loc)
                conf = float(ner_loc.get("confidence") or 0.0)
                if cand and conf >= NER_CONF_THRESHOLD:
                    return cand, {
                        "method": "resume_ner",
                        "score": 85,
                        "confidence": conf,
                        "reasons": ["pipeline_stage=2", "ner_conf_ok"]
                    }
        except Exception:
            pass

    # 2b) Gazetteer/heuristic inference (regex + scoring)
    loc2, meta2 = infer_location_from_resume_details(full, dom_city_only=dom_city_only)
    if loc2:
        meta2 = dict(meta2 or {})
        meta2["method"] = meta2.get("method") or "resume_infer"
        meta2["reasons"] = list(meta2.get("reasons") or []) + ["pipeline_stage=2"]
        return loc2, meta2

    # ---- Stage 3: Phone country fallback ----
    c = _country_from_phone(phone)
    if c:
        return c, {"method": "phone_country_fallback", "score": 40, "confidence": 0.40, "reasons": ["pipeline_stage=3"]}

    return "", None


async def extract_location_hybrid(pdf_text: str, dom_city_only: Optional[str] = None):
    """Try NER first, fallback to regex"""
    if USE_NER and _NER_EXTRACTOR:
        try:
            loop = asyncio.get_running_loop()
            ner_result = await loop.run_in_executor(
                GLOBAL_PDF_EXECUTOR,
                lambda: _NER_EXTRACTOR.extract_from_resume(pdf_text, debug=NER_DEBUG)
            )
            if ner_result and ner_result.confidence >= NER_MIN_CONFIDENCE:
                location_str = format_location(ner_result)
                metadata = {
                    "method": "ner",
                    "confidence": ner_result.confidence,
                    "score": int(ner_result.confidence * 100),
                    "reasons": [f"ner_{ner_result.method}"]
                }
                return location_str, metadata
        except Exception as e:
            if NER_DEBUG:
                print(f"   NER error: {e}")
    
    # Fallback to regex
    return infer_location_from_resume_details(pdf_text, dom_city_only)

# ─── Keywords matching ──────────────────────────────────────────────────────
def _normalize_for_kw(s: str) -> str:
    """Normalize text for robust keyword matching (PDFs often contain NBSP/newlines/zero-width chars)."""
    s = unicodedata.normalize("NFKC", s or "")

    # Remove invisible "zero-width" characters that can break matching in PDFs
    s = re.sub(r"[\u200B-\u200D\uFEFF]", "", s)  # ZWSP/ZWNJ/ZWJ/BOM

    # Normalize common weird spaces found in PDFs
    s = s.replace("\u00A0", " ")  # NBSP
    s = s.replace("\u202F", " ")  # narrow NBSP
    s = s.replace("\u2007", " ")  # figure space

    # Collapse all whitespace (spaces/newlines/tabs) to single spaces
    s = re.sub(r"\s+", " ", s)
    return s.strip()

def find_keyword_hits(
    texts: List[str],
    keywords: List[str],
    case_insensitive: bool = True,
    whole_word: bool = True,
    max_items: int = 50,
) -> List[str]:
    """
    Ctrl+F style keyword search across:
      - name/title/company/location
      - screening Q/A
      - PDF text (and OCR output if used)

    Handles:
      - ISO\n45001 style line breaks
      - weird PDF spaces
      - zero-width characters

    Improvements:
      - Separator-flexible matching: "ISO45001", "ISO-45001", "ISO 45001" all match "ISO 45001"
      - ISO list heuristic: "ISO Standard (9001, 45001, 22000)" will still match "ISO 45001"
    """

    # One normalized haystack to search in
    hay = _normalize_for_kw(" ".join(t for t in texts if t) or "")
    if not hay:
        return []

    flags = re.I if case_insensitive else 0
    found: List[Tuple[str, int]] = []

    def _kw_display_key(kw: str) -> str:
        kw = (kw or "").strip()
        # treat "(ISO 45001)" the same as "ISO 45001" for matching
        if kw.startswith("(") and kw.endswith(")"):
            inner = kw[1:-1].strip()
            if inner:
                kw = inner
        return kw

    def _sep_flexible_pattern(kw_norm: str) -> str:
        # Build something like: ISO[\W_]*45001 (allows space/dash/slash/underscore or nothing)
        tokens = re.findall(r"[A-Za-z]+|\d+", kw_norm)
        if not tokens:
            return ""
        mid = r"[\W_]*"
        body = mid.join(re.escape(t) for t in tokens if t)
        if not body:
            return ""
        if whole_word:
            return rf"(?<![A-Za-z0-9]){body}(?![A-Za-z0-9])"
        return body

    def _iso_list_hit(num: str) -> bool:
        # True if a "line-ish" chunk contains ISO and the bare number, e.g. "ISO Standard (9001, 45001, 22000)"
        if not num:
            return False
        num_rx = re.compile(rf"(?<!\d){re.escape(num)}(?!\d)", flags)
        # Keep chunking simple; hay already has whitespace collapsed
        chunks = re.split(r"[\n•]+", hay)
        for ch in chunks:
            if re.search(r"\bISO\b", ch, flags) and num_rx.search(ch):
                return True
        return False

    for kw in keywords:
        kw_disp = (kw or "").strip()
        if not kw_disp:
            continue

        # normalized keyword for robust tokenization
        kw_key = _kw_display_key(kw_disp)
        kw_norm = _normalize_for_kw(kw_key)
        if not kw_norm:
            continue

        pat = _sep_flexible_pattern(kw_norm)
        count = 0

        if pat:
            rx = re.compile(pat, flags)
            count = len(list(rx.finditer(hay)))

        # ISO heuristic: if keyword is ISO <digits>, allow matching the digits when clearly in an ISO context
        if count == 0:
            tokens = re.findall(r"[A-Za-z]+|\d+", kw_norm)
            if tokens and tokens[0].lower() == "iso":
                num = next((t for t in tokens[1:] if t.isdigit()), "")
                if num and _iso_list_hit(num):
                    count = 1

        if count > 0:
            found.append((kw_disp, count))

    # Dedupe by lowercase keyword; keep first display form, add (xN) if > 1
    out: List[str] = []
    seen_lower = set()
    for kw_disp, count in found:
        key = kw_disp.lower()
        if key in seen_lower:
            continue
        seen_lower.add(key)
        out.append(kw_disp if count <= 1 else f"{kw_disp} (x{count})")
        if len(out) >= max_items:
            break

    return out

# ─── Screening extraction ───────────────────────────────────────────────────
_Q_TRAIL_PUNCT = re.compile(r"[:?\s]+$")
_WS_MULTI = re.compile(r"\s+")
CERT_START = re.compile(r"^i\s+certify\s+that\s+to\s+the\s+best\s+of\s+my\s+knowledge", re.I)

def canonical_q(q: str) -> str:
    q = _Q_TRAIL_PUNCT.sub("", q.strip())
    q = _WS_MULTI.sub(" ", q)
    return q.lower()

def headerize(q: str) -> str:
    q = _Q_TRAIL_PUNCT.sub("", q.strip())
    q = _WS_MULTI.sub(" ", q)
    return q

SCREENING_TAB_SELECTOR = "spl-typography-title:has-text('Screening')"

async def click_screening_tab(page) -> bool:
    try:
        await page.locator(SCREENING_TAB_SELECTOR).first.click()
        await page.wait_for_timeout(500)
        return True
    except Exception:
        try:
            await page.get_by_role("tab", name=re.compile(r"^Screening$", re.I)).click(timeout=1500)
            await page.wait_for_timeout(500)
            return True
        except Exception:
            return True


RESUME_TAB_SELECTOR = "spl-typography-title:has-text('Resume')"

async def click_resume_tab(page) -> bool:
    """Open the Resume view so the embedded viewer/link is present (matches what you Ctrl+F manually)."""
    try:
        # Prefer the tab role if present
        await page.get_by_role("tab", name=re.compile(r"^Resume$", re.I)).click(timeout=2000)
        await page.wait_for_timeout(600)
        return True
    except Exception:
        pass
    try:
        await page.locator(RESUME_TAB_SELECTOR).first.click(timeout=2000)
        await page.wait_for_timeout(600)
        return True
    except Exception:
        return False

async def get_screening_pairs(page):
    try:
        h = await page.locator("text=Screening Questions").first.element_handle()
        if h:
            root = await h.evaluate_handle(
                "el => el.closest('section') || el.closest('spl-card') || el.parentElement || el"
            )
        else:
            root = await page.locator("section:has-text('Screening')").first.element_handle()
            if not root:
                return []
    except Exception:
        return []

    texts = await root.evaluate(r"""
      (el) => {
        const w = document.createTreeWalker(el, NodeFilter.SHOW_TEXT, null);
        const out = [];
        while (w.nextNode()) {
          const t = (w.currentNode.nodeValue || '').replace(/\s+/g,' ').trim();
          if (!t) continue;
          if (/^Last edit was made/i.test(t)) continue;
          if (/^See versions$/i.test(t)) continue;
          if (/^Screening Questions$/i.test(t)) continue;
          out.push(t);
        }
        return out;
      }
    """)

    Q_RE = re.compile(
        r"(?ix)^(please|specify|provide|enter|select|choose|state|list|give|what|which|when|where|how|are\s+you|do\s+you|did\s+you|have\s+you|i\s+certify)"
    )

    def looks_like_q(s):
        s = s.strip()
        return s.endswith(":") or s.endswith("?") or bool(Q_RE.match(s))

    pairs, q, ans = [], None, []
    for t in texts:
        if CERT_START.search(t):
            if q:
                pairs.append([q, " ".join(ans).strip()])
            q, ans = None, []
            continue
        if looks_like_q(t):
            if q:
                pairs.append([q, " ".join(ans).strip()])
            q, ans = t, []
        else:
            if q:
                ans.append(t)
    if q:
        pairs.append([q, " ".join(ans).strip()])

    out = []
    for q0, a0 in pairs:
        if CERT_START.search(q0):
            continue
        out.append([headerize(q0), _WS_MULTI.sub(" ", (a0 or "").strip())])
    return out

# ─── Link collector ─────────────────────────────────────────────────────────
PROFILE_LINK_SELECTOR = (
    'a[href^="/app/people/applications/"], a[href^="/app/people/profile/"], '
    'sr-link[href^="/app/people/applications/"], sr-link[href^="/app/people/profile/"]'
)
LINK_COLLECT_TIMEOUT = 45000  # Increased from 20s; SmartRecruiters can load slowly

async def dismiss_overlays(page):
    try:
        await page.keyboard.press("Escape")
    except Exception:
        pass
    candidates = [
        page.get_by_role("button", name=re.compile(r"^(continue|accept|got it|agree|ok|okay|close)$", re.I)),
        page.locator("button:has-text('Continue')"),
        page.locator("button:has-text('Accept')"),
        page.locator("button:has-text('Got it')"),
        page.locator("[role=button]:has-text('Continue')"),
        page.locator("button[aria-label='Close'], [aria-label='Close']")
    ]
    for loc in candidates:
        try:
            n = await loc.count()
            for i in range(n):
                try:
                    el = loc.nth(i)
                    await el.click(timeout=300)
                    await page.wait_for_timeout(120)
                except Exception:
                    pass
        except Exception:
            pass

async def collect_all_profile_links(page) -> List[str]:
    url = page.url
    print(f"   📍 Current page: {url[:100]}{'...' if len(url) > 100 else ''}")
    if "smartrecruiters.com" not in url.lower():
        print("   ⚠️  Not on SmartRecruiters. Ensure the first tab is the prospect/applicant list.")
    await page.wait_for_selector(PROFILE_LINK_SELECTOR, timeout=LINK_COLLECT_TIMEOUT)
    await dismiss_overlays(page)

    paths: List[str] = await page.evaluate(r"""
(() => new Promise((resolve) => {
  const normalizePath = (raw) => {
    try {
      const u = new URL(raw, location.origin);
      const m = u.pathname.match(/^\/app\/people\/(?:applications|profile)\/[0-9A-Fa-f\-]+/);
      return m ? (m[0] + "/") : (u.pathname || "").replace(/\/?$/, "/");
    } catch { return ""; }
  };
  const textAll = (root=document.body) => (root.innerText || root.textContent || "");
  const parseExpectedTotal = () => {
    const txt = textAll().replace(/,/g, " ");
    const m = txt.match(/Showing\s+\d+\s+of\s+(\d+)\s+(applicants|prospects|people)/i);
    return m ? parseInt(m[1], 10) : null;
  };
  const style = (el) => (el ? getComputedStyle(el) : null);
  const isScrollable = (el) => !!el && (el.scrollHeight - el.clientHeight > 8) &&
    /(auto|scroll)/i.test((style(el).overflowY || style(el).overflow || ''));
  const scroller = (() => {
    const candidates = [
      'div[role="grid"]',
      'section:has(div[role="grid"])',
      'div[aria-label*="Applicants"]',
      'div.spl-scroll-y',
      'div.spl-scroll-container'
    ];
    for (const sel of candidates) {
      const el = document.querySelector(sel);
      if (isScrollable(el)) return el;
    }
    const linkSel = 'a[href^="/app/people/applications/"], a[href^="/app/people/profile/"],' +
                    'sr-link[href^="/app/people/applications/"], sr-link[href^="/app/people/profile/"]';
    let el = document.querySelector(linkSel);
    while (el && el !== document.body) {
      if (isScrollable(el)) return el;
      el = el.parentElement;
    }
    return document.scrollingElement || document.documentElement;
  })();
  const clickLoadMoreIfAny = () => {
    const all = Array.from(document.querySelectorAll("button, [role='button']"));
    for (const b of all) {
      const t = (b.innerText || b.textContent || "").trim().toLowerCase();
      if (!t) continue;
      if (t === "load more" || t.startsWith("load")) {
        if (!b.disabled) { try { b.click(); return true; } catch {} }
      }
    }
    return false;
  };
  const linkSel = 'a[href^="/app/people/applications/"], a[href^="/app/people/profile/"],' +
                  'sr-link[href^="/app/people/applications/"], sr-link[href^="/app/people/profile/"]';
  const seen = new Set();
  let expectedTotal = parseExpectedTotal();
  const MAX_RUN_MS = 3 * 60 * 1000;
  const IDLE_NO_GROW_MS = 2500;
  const CONFIRM_STABLE_MS = 900;
  const LOAD_MORE_COOLDOWN_MS = 600;
  let lastAddTime = performance.now();
  let startTime = performance.now();
  let reachedExpectedAt = null;
  let lastLoadMore = 0;
  const harvest = () => {
    const nodes = document.querySelectorAll(linkSel);
    let added = 0;
    for (const n of nodes) {
      const raw = n.getAttribute("href") || n.href || "";
      const norm = normalizePath(raw);
      if (norm && !seen.has(norm)) { seen.add(norm); added++; }
    }
    if (added > 0) lastAddTime = performance.now();
    return added;
  };
  harvest();
  const step = () => {
    const before = scroller.scrollTop;
    scroller.scrollTop = Math.min(scroller.scrollTop + scroller.clientHeight * 1.6, scroller.scrollHeight);
    if (scroller.scrollTop === before) {
      scroller.scrollTop = scroller.scrollHeight - scroller.clientHeight - 1;
    }
    const now = performance.now();
    if ((now - lastAddTime) > 400 && (now - lastLoadMore) > LOAD_MORE_COOLDOWN_MS) {
      if (clickLoadMoreIfAny()) lastLoadMore = now;
    }
    harvest();
    if (expectedTotal && seen.size >= expectedTotal) {
      if (reachedExpectedAt === null) reachedExpectedAt = now;
      if (now - reachedExpectedAt >= CONFIRM_STABLE_MS) return resolve(Array.from(seen));
    }
    if (!expectedTotal) {
      if (now - lastAddTime >= IDLE_NO_GROW_MS) return resolve(Array.from(seen));
    }
    if (now - startTime >= MAX_RUN_MS) return resolve(Array.from(seen));
    requestAnimationFrame(step);
  };
  requestAnimationFrame(step);
}))""")
    return paths

# ─── Job details (for report filename) ───────────────────────────────────────
JOB_ROLE_NAME_SELECTORS = [
    "app-job-title-section spl-typography-title",
    "#st-jobDetailsPage app-job-title-section spl-typography-title",
]
JOB_ROLE_NUMBER_SELECTORS = [
    "app-job-attribute:nth-child(3) spl-typography-body",
    "#st-jobDetailsPage app-job-attribute:nth-child(3) spl-typography-body",
]

def _sanitize_filename(s: str) -> str:
    """Make string safe for use in filenames."""
    if not s:
        return ""
    s = re.sub(r'[/\\:*?"<>|]', "", s)
    s = re.sub(r"\s+", "_", s.strip())
    return s[:80] if s else ""

async def get_job_details(page) -> Tuple[str, str]:
    """
    Extract role name and requisition number from the job details page.
    Returns (role_name, role_number). Empty strings if not found.
    """
    role_name, role_number = "", ""
    for sel in JOB_ROLE_NAME_SELECTORS:
        try:
            el = await page.query_selector(sel)
            if el:
                role_name = _clean(await el.inner_text()) or ""
                if role_name:
                    break
        except Exception:
            continue
    for sel in JOB_ROLE_NUMBER_SELECTORS:
        try:
            el = await page.query_selector(sel)
            if el:
                raw = _clean(await el.inner_text()) or ""
                m = re.search(r"([Rr]\d+_\d+)", raw)
                role_number = m.group(1) if m else raw.split("|")[0].strip() if "|" in raw else raw
                if role_number:
                    break
        except Exception:
            continue
    return role_name, role_number

# ─── DOM helpers ────────────────────────────────────────────────────────────
PROFILE_LOCATION_SELECTOR = (
    "#st-candidateView > sr-page-wrapper > div > div > div > sr-candidate-personal-info > "
    "div > div > div > div.spl-flex.spl-gap-2.spl-flex-wrap.spl-flex-col.sm\\:spl-flex-row > "
    "div:nth-child(1) > spl-typography-body"
)
NAME_SELECTOR = "#st-applicantName > spl-truncate"
TITLE_COMPANY_SELECTOR = (
    "#st-candidateView > sr-page-wrapper > div > div > div > "
    "sr-candidate-personal-info > div > div > div > "
    "div.spl-flex.spl-flex-col.spl-gap-1 > spl-typography-body > spl-truncate"
)

async def get_profile_dom_location(page) -> str:
    def _sanitize_dom_loc(val: str) -> str:
        s = strip_contacts_and_noise(val)
        if not s or re.search(r"\+\d|\d{5,}", s):
            return ""
        tokens = [t.strip() for t in re.split(r"[,\u2013|/;-]+", s) if t.strip()]
        if len(tokens) < 2:
            return ""
        country = _country_lookup(tokens[-1])
        if not country:
            return ""
        left = ", ".join(tokens[:-1]).strip(" ,")
        if not left:
            return country["name"]
        last_left = tokens[-2].strip()
        if len(last_left) < 3 or last_left.lower() in _STOPWORDS:
            return ""
        return f"{left}, {country['name']}"

    try:
        el = await page.wait_for_selector(PROFILE_LOCATION_SELECTOR, timeout=3000)
        val = await el.inner_text()
        norm = _sanitize_dom_loc(_clean(val))
        if norm:
            return norm
    except Exception:
        pass
    try:
        texts = await page.eval_on_selector_all(
            "sr-candidate-personal-info spl-typography-body",
            "els => els.map(e => (e.innerText || '').trim()).filter(Boolean)"
        )
        for t in texts:
            norm = _sanitize_dom_loc(t)
            if norm:
                return norm
    except Exception:
        pass
    return ""

async def get_dom_city_only(page) -> str:
    try:
        texts = await page.eval_on_selector_all(
            "sr-candidate-personal-info spl-typography-body",
            "els => els.map(e => (e.innerText || '').trim()).filter(Boolean)"
        )
    except Exception:
        return ""
    for t in texts:
        s = strip_contacts_and_noise(t)
        toks = [x.strip() for x in re.split(r"[,\u2013|/;-]+", s) if x.strip()]
        if not toks:
            continue
        if _country_lookup(toks[-1]):
            continue
        for tok in toks:
            if len(tok) >= 3 and tok.lower() not in _STOPWORDS and _is_real_city(tok):
                return tok
    return ""

async def get_candidate_name(page) -> str:
    try:
        el = await page.query_selector(NAME_SELECTOR)
        return _clean(await el.inner_text()) if el else ""
    except Exception:
        return ""

async def get_title_and_company(page):
    try:
        el = await page.query_selector(TITLE_COMPANY_SELECTOR)
        raw = _clean(await el.inner_text()) if el else ""
        parts = raw.split(" at ", 1)
        return (parts[0].strip(), parts[1].strip()) if len(parts) == 2 else (raw, "")
    except Exception:
        return "", ""





async def get_dom_phone(page) -> str:
    """Best-effort phone scrape from the profile DOM (if present)."""
    # Prefer an explicit tel: link if the UI exposes it.
    try:
        tel = await page.get_attribute('a[href^="tel:"]', "href")
        if tel:
            phone = tel.replace("tel:", "").strip()
            return _clean(phone)
    except Exception:
        pass

    # Fallback: look for a +<digits> style phone in small header areas (low risk, avoids scanning whole resume text).
    try:
        # This searches visible text nodes matching a basic E.164-ish pattern.
        locator = page.locator("text=/\+\d[\d\s\-()]{7,}/").first
        if await locator.count() > 0:
            txt = _clean(await locator.inner_text())
            m = re.search(r"(\+\d[\d\s\-()]{7,})", txt)
            if m:
                return _clean(m.group(1))
    except Exception:
        pass

    return ""


async def get_dom_resume_text(page, max_chars: int = 40000) -> str:
    """
    Grab as much human-visible text as possible from the candidate view.
    This is the closest equivalent to "Ctrl+F on the resume" when SmartRecruiters renders the resume in the DOM.
    """
    try:
        txt = await page.evaluate(r"""
() => {
  const root = document.querySelector('#st-candidateView') || document.body;

  // Try resume-ish containers first (varies by SR tenant/version)
  const selectors = [
    'sr-resume-viewer',
    'sr-candidate-resume',
    'sr-resume',
    '[data-testid*="resume"]',
    '[class*="resume"]',
    '[id*="resume"]'
  ];

  for (const sel of selectors) {
    const el = root.querySelector(sel);
    if (!el) continue;
    const t = (el.innerText || el.textContent || '').replace(/\s+/g, ' ').trim();
    if (t && t.length > 200) return t;
  }

  // Fallback: everything visible inside candidate view
  return (root.innerText || root.textContent || '').trim();
}
        """)
        if not txt:
            return ""
        txt = txt.strip()
        return txt[:max_chars]
    except Exception:
        return ""

# ─── PDF ────────────────────────────────────────────────────────────────────
GLOBAL_PDF_EXECUTOR: Optional[concurrent.futures.ThreadPoolExecutor] = None
PDF_URL_HINTS = ("/app/people/api/attachments/", ".pdf")

# ─── OCR Functions ──────────────────────────────────────────────────────────

def get_pdf_hash(pdf_bytes: bytes) -> str:
    """Generate hash of PDF for caching"""
    return hashlib.md5(pdf_bytes).hexdigest()

def get_cached_ocr_text(pdf_hash: str) -> Optional[str]:
    """Retrieve cached OCR text if available"""
    if not USE_OCR:
        return None
    cache_file = OCR_CACHE_DIR / f"{pdf_hash}.txt"
    if cache_file.exists():
        try:
            return cache_file.read_text(encoding='utf-8')
        except Exception:
            return None
    return None

def save_ocr_cache(pdf_hash: str, text: str):
    """Save OCR text to cache"""
    if not USE_OCR:
        return
    try:
        cache_file = OCR_CACHE_DIR / f"{pdf_hash}.txt"
        cache_file.write_text(text, encoding='utf-8')
    except Exception:
        pass

def ocr_pdf_bytes(pdf_bytes: bytes, max_pages: int = OCR_MAX_PAGES) -> str:
    """
    Extract text from PDF using OCR.
    Only processes first N pages for speed.
    """
    if not USE_OCR or not PADDLEOCR_AVAILABLE:
        return ""
    
    # Check cache first
    pdf_hash = get_pdf_hash(pdf_bytes)
    cached_text = get_cached_ocr_text(pdf_hash)
    if cached_text is not None:
        if OCR_DEBUG:
            print(f"   💾 OCR cache hit ({len(cached_text)} chars)")
        return cached_text
    
    try:
        ocr = get_ocr_engine()
        if not ocr:
            return ""
        
        # Open PDF
        pdf = fitz.open(stream=bytes(pdf_bytes), filetype="pdf")
        ocr_text = ""
        pages_processed = 0
        
        # Process only first N pages
        for page_num in range(min(max_pages, pdf.page_count)):
            try:
                page = pdf.load_page(page_num)
                
                # Convert page to image
                pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))  # 2x zoom for better OCR
                img_bytes = pix.tobytes("png")
                img = Image.open(io.BytesIO(img_bytes))
                
                # Run OCR (accepts bytes, path, or numpy array)
                result = ocr.ocr(img_bytes, cls=True)
                
                # Extract text — robust parsing for PaddleOCR 2.x and 3.x
                # Format: [[[box, (text, conf)], ...], None] or list of pages
                lines = []
                if result and isinstance(result, (list, tuple)):
                    for page in result:
                        if page is None:
                            continue
                        if isinstance(page, (list, tuple)):
                            for item in page:
                                if isinstance(item, (list, tuple)) and len(item) >= 2:
                                    lines.append(item)
                for line in lines:
                    if not line or len(line) < 2:
                        continue
                    part = line[1]
                    if isinstance(part, (tuple, list)) and len(part) >= 1:
                        text_content = part[0] if isinstance(part[0], str) else str(part[0])
                    else:
                        text_content = str(part)
                    if text_content.strip():
                        ocr_text += text_content + "\n"
                
                pages_processed += 1
                
            except Exception as e:
                if OCR_DEBUG:
                    print(f"   ⚠️  OCR page {page_num} error: {type(e).__name__}")
                continue
        
        pdf.close()
        
        # Save to cache
        if ocr_text:
            save_ocr_cache(pdf_hash, ocr_text)
            if OCR_DEBUG:
                print(f"   ✅ OCR extracted {len(ocr_text)} chars from {pages_processed} pages")
        
        return ocr_text
        
    except Exception as e:
        if OCR_DEBUG:
            print(f"   ⚠️  OCR failed: {type(e).__name__}: {str(e)[:100]}")
        return ""

async def close_stray_pdf_tabs(ctx):
    for pg in list(ctx.pages):
        try:
            u = (pg.url or "").lower()
        except Exception:
            continue
        if any(h in u for h in PDF_URL_HINTS):
            try:
                await pg.close()
            except Exception:
                pass

async def open_resume_and_get_text(ctx, page) -> str:
    """
    Resume/PDF extractor (best path first, no new tabs):
      1) If resume is embedded (iframe/embed/object), grab its URL and download PDF bytes
      2) Else try "Latest Resume" link (top of profile or attachments)
      3) Else click a resume/attachment row and capture the PDF network response

    Returns extracted PDF text (PyMuPDF) + OCR fallback when applicable.
    """

    if LOCATION_DEBUG:
        print("   🔍 PDF extraction v2: start")

    def extract_text_from_pdf_bytes(pdf_bytes: bytes) -> str:
        text_out = ""
        pdf_open_failed = False
        try:
            pdf = fitz.open(stream=pdf_bytes, filetype="pdf")
            if LOCATION_DEBUG:
                print(f"   🔍 PDF opened: {pdf.page_count} pages")
            for i in range(min(PDF_MAX_PAGES, pdf.page_count)):
                text_out += pdf.load_page(i).get_text("text") or ""
            pdf.close()
        except Exception as e:
            pdf_open_failed = True
            if LOCATION_DEBUG:
                print(f"   ⚠️  PDF open/extract failed: {type(e).__name__}: {str(e)[:120]}")

        # OCR fallback for scanned / low-text PDFs
        ocr_text = ""
        if (len(text_out) < OCR_MIN_TEXT_LENGTH or pdf_open_failed) and USE_OCR:
            if LOCATION_DEBUG:
                print(f"   ⚠️  OCR fallback triggered (len={len(text_out)})")
            ocr_text = ocr_pdf_bytes(pdf_bytes, OCR_MAX_PAGES)
            if ocr_text:
                return (text_out + "\n" + ocr_text).strip()

        # Debug: always run OCR on first page to verify it works (adds ~2–5s per profile)
        if OCR_DEBUG_ALWAYS_FIRST_PAGE and USE_OCR and pdf_bytes:
            if LOCATION_DEBUG:
                print(f"   🔬 OCR_DEBUG_ALWAYS_FIRST_PAGE: running OCR on page 0")
            debug_ocr = ocr_pdf_bytes(pdf_bytes, max_pages=1)
            if debug_ocr:
                text_out = (text_out + "\n" + debug_ocr).strip()
                if LOCATION_DEBUG:
                    print(f"   🔬 OCR page 0 extracted {len(debug_ocr)} chars")

        return (text_out or "").strip()

    async def fetch_bytes(url: str):
        if not url:
            return None
        try:
            r = await ctx.request.get(url, timeout=20000)
            if not r.ok:
                if LOCATION_DEBUG:
                    print(f"   ⚠️  GET failed {r.status} url={url[:120]}")
                return None
            b = await r.body()
            if not b or len(b) < 800:
                if LOCATION_DEBUG:
                    print(f"   ⚠️  GET returned tiny body ({len(b) if b else 0} bytes)")
                return None
            return b
        except Exception as e:
            if LOCATION_DEBUG:
                print(f"   ⚠️  GET error: {type(e).__name__}: {str(e)[:120]}")
            return None

    def looks_like_pdfish(u: str) -> bool:
        if not u:
            return False
        ul = u.lower()
        return (".pdf" in ul) or ("application/pdf" in ul) or ("/app/people/api/attachments/" in ul) or ("/attachments/" in ul)

    # ── 1) Embedded viewer URL (iframe/embed/object) ───────────────────────
    try:
        viewer_url = await page.evaluate(r"""
() => {
  const root = document.querySelector('#st-candidateView') || document;

  const getAttr = (el, attr) => {
    try { return (el && el.getAttribute && el.getAttribute(attr)) || ''; } catch { return ''; }
  };

  // Common patterns
  const iframe = root.querySelector('iframe[src]');
  if (iframe) return getAttr(iframe, 'src');

  const embed = root.querySelector('embed[src]');
  if (embed) return getAttr(embed, 'src');

  const obj = root.querySelector('object[data]');
  if (obj) return getAttr(obj, 'data');

  return '';
}
        """)
        if viewer_url and looks_like_pdfish(viewer_url):
            if LOCATION_DEBUG:
                print(f"   ✅ viewer URL found: {viewer_url[:140]}")
            pdf_bytes = await fetch_bytes(viewer_url)
            if pdf_bytes:
                loop = asyncio.get_running_loop()
                return await loop.run_in_executor(GLOBAL_PDF_EXECUTOR, extract_text_from_pdf_bytes, pdf_bytes)
    except Exception as e:
        if LOCATION_DEBUG:
            print(f"   ⚠️  viewer URL path failed: {type(e).__name__}: {str(e)[:120]}")

    # ── 2) Latest Resume link (top of profile) ─────────────────────────────
    try:
        latest = page.get_by_role("link", name=re.compile(r"latest\s+resume", re.I))
        if await latest.count() > 0:
            href = await latest.first.get_attribute("href")
            if href and looks_like_pdfish(href):
                if LOCATION_DEBUG:
                    print(f"   ✅ Latest Resume href found: {href[:140]}")
                pdf_bytes = await fetch_bytes(href)
                if pdf_bytes:
                    loop = asyncio.get_running_loop()
                    return await loop.run_in_executor(GLOBAL_PDF_EXECUTOR, extract_text_from_pdf_bytes, pdf_bytes)
    except Exception:
        pass

    # ── 2b) Any obvious attachments href in DOM (works in some tenants) ────
    try:
        href = await page.evaluate(r"""
(() => {
  const root = document.querySelector('sr-attachments-v2') || document;
  const a = root.querySelector('a[href*="/app/people/api/attachments/"]') || root.querySelector('a[href$=".pdf"]');
  return (a && (a.href || a.getAttribute('href'))) ? (a.href || a.getAttribute('href')) : '';
})()
        """)
        if href and looks_like_pdfish(href):
            if LOCATION_DEBUG:
                print(f"   ✅ Attachments href found: {href[:140]}")
            pdf_bytes = await fetch_bytes(href)
            if pdf_bytes:
                loop = asyncio.get_running_loop()
                return await loop.run_in_executor(GLOBAL_PDF_EXECUTOR, extract_text_from_pdf_bytes, pdf_bytes)
    except Exception:
        pass

    # ── 3) Fallback: click + capture PDF response ──────────────────────────
    if LOCATION_DEBUG:
        print("   ↩️  fallback: click + expect_response")

    CLICK_CANDIDATES = [
        # Your original brittle selector first (may still work)
        "#st-candidateView > sr-page-wrapper > div > aside > sr-job-application-sidebar > div > sr-attachments-v2",
        # More general
        "sr-attachments-v2 sr-attachment-row spl-link-button",
        "sr-attachments-v2 sr-attachment-row",
        "text=/latest resume|resume|cv|attachment/i",
    ]

    async def _click_any():
        last_err = None
        for sel in CLICK_CANDIDATES:
            try:
                loc = page.locator(sel).first
                if await loc.count() == 0:
                    continue
                await loc.scroll_into_view_if_needed(timeout=1500)
                await loc.click(timeout=2500, force=True)
                return True
            except Exception as e:
                last_err = e
                continue
        if LOCATION_DEBUG and last_err:
            print(f"   ⚠️  click failed: {type(last_err).__name__}: {str(last_err)[:120]}")
        return False

    try:
        def _is_pdf(resp):
            u = (resp.url or "").lower()
            ct = (resp.headers.get("content-type", "") or "").lower()
            return ("pdf" in ct) or u.endswith(".pdf") or ("/app/people/api/attachments/" in u) or ("/attachments/" in u)

        async with page.expect_response(_is_pdf, timeout=15000) as resp_info:
            clicked = await _click_any()
            if not clicked:
                return ""

        resp = await resp_info.value
        pdf_bytes = await resp.body()
        if pdf_bytes and len(pdf_bytes) > 800:
            loop = asyncio.get_running_loop()
            return await loop.run_in_executor(GLOBAL_PDF_EXECUTOR, extract_text_from_pdf_bytes, pdf_bytes)

    except Exception as e:
        if LOCATION_DEBUG:
            print(f"   ⚠️  fallback capture failed: {type(e).__name__}: {str(e)[:150]}")

    return ""


# ─── Context perf tuning ────────────────────────────────────────────────────
TRACKER_HOST_SNIPPETS = (
    "google-analytics.com", "googletagmanager.com", "doubleclick.net",
    "facebook.com", "fbcdn.net", "appsflyer", "hotjar", "segment.io",
    "sentry.io", "amplitude.com", "mixpanel.com"
)
BLOCKED_EXTS = (
    ".jpg", ".jpeg", ".png", ".gif", ".webp", ".svg",
    ".mp4", ".mp3", ".woff", ".woff2", ".ttf", ".otf", ".ico", ".map"
)
BLOCKED_RESOURCE_TYPES = frozenset([
    "image", "media", "font", "websocket",
    "manifest", "texttrack", "eventsource"
])

def is_tracker(url: str) -> bool:
    h = urlparse(url).hostname or ""
    return any(p in h for p in TRACKER_HOST_SNIPPETS)

def is_heavy_asset(url: str) -> bool:
    u = url.lower()
    return any(u.endswith(ext) for ext in BLOCKED_EXTS)

async def setup_context_perf_tuning(ctx):
    async def handler(route, request):
        rtype = request.resource_type
        url = request.url
        if rtype in BLOCKED_RESOURCE_TYPES or is_heavy_asset(url) or is_tracker(url):
            return await route.abort()
        return await route.continue_()

    await ctx.route("**/*", handler)
    try:
        await ctx.set_extra_http_headers({
            'Accept-Language': 'en-US,en;q=0.9',
            'Accept-Encoding': 'gzip, deflate, br',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Cache-Control': 'max-age=0'
        })
    except Exception:
        pass

async def ensure_page_viewport(page):
    try:
        await page.set_viewport_size(VIEWPORT)
    except Exception:
        pass

# ─── Navigation pacer ───────────────────────────────────────────────────────
class NavPacer:
    def __init__(self, min_gap_ms: int):
        self.min_gap = max(120, min_gap_ms) / 1000.0
        self._lock = asyncio.Lock()
        self._last = 0.0

    async def wait_turn(self):
        async with self._lock:
            now = time.perf_counter()
            gap = now - self._last
            if gap < self.min_gap:
                await asyncio.sleep(self.min_gap - gap + random.uniform(0.02, 0.06))
            self._last = time.perf_counter()

# ─── Worker loop ────────────────────────────────────────────────────────────
async def worker_loop(wid: int, ctx, page, q: asyncio.Queue, pacer: NavPacer, results: dict):
    processed = 0
    page.set_default_timeout(10000)
    page.on("dialog", lambda d: asyncio.create_task(d.dismiss()))
    await ensure_page_viewport(page)

    while True:
        item = await q.get()
        if item is None:
            q.task_done()
            break
        index, total, path = item
        url = BASE_URL + path
        print(f"➡️ [W{wid} {index}/{total}] {url}")
        await pacer.wait_turn()
        await asyncio.sleep(random.uniform(HUMAN_DELAY_MIN, HUMAN_DELAY_MAX))

        for attempt in range(1, RETRY_LIMIT + 1):
            try:
                await page.goto(url, timeout=PAGE_LOAD_TIMEOUT, wait_until="networkidle")
                await dismiss_overlays(page)
                await page.wait_for_timeout(400)
                await page.wait_for_selector(NAME_SELECTOR, timeout=6000)

                name = await get_candidate_name(page)
                title, company = await get_title_and_company(page)

                # Ensure Resume view is open (matches manual Ctrl+F behaviour)
                await click_resume_tab(page)

                dom_resume_text = await get_dom_resume_text(page)
                if LOCATION_DEBUG:
                    print(f"   🧾 DOM resume text: {len(dom_resume_text)} chars" if dom_resume_text else "   🧾 DOM resume text: (none)")
                # Extract resume PDF text once (used for keywords + location fallback)
                pdf_text = await open_resume_and_get_text(ctx, page)

                # Location extraction — DOM first, resume fallback only if missing
                location = ""
                loc_debug = None

                dom_location = await get_profile_dom_location(page)
                if LOCATION_DEBUG:
                    print(f"   📄 PDF: {len(pdf_text)} chars" if pdf_text else "   ⚠️  PDF: No text")
                    if dom_location:
                        print(f"   🌐 DOM location: {dom_location}")

                if dom_location:
                    location = dom_location
                    loc_debug = {
                        "method": "dom",
                        "confidence": 0.95,
                        "score": 95,
                        "reasons": ["dom_selector"]
                    }
                else:
                    dom_city_only = await get_dom_city_only(page)
                    phone = await get_dom_phone(page)

                    print(f"   📭 DOM missing → resume inference used (phone={'yes' if phone else 'no'})")

                    # 1) Infer from DOM resume text (fast)
                    location, loc_debug = await infer_location_from_resume(
                        dom_resume_text or "",
                        phone=phone,
                        dom_city_only=dom_city_only
                    )

                    # 2) If still missing, try PDF text (slower / edge cases)
                    if not location and pdf_text:
                        location, loc_debug = await infer_location_from_resume(
                            pdf_text,
                            phone=phone,
                            dom_city_only=dom_city_only
                        )

                    if location:
                        print(f"   🧭 DOM missing → resume inference used → {location}")
                    else:
                        loc_debug = {
                            "method": "none",
                            "confidence": 0.0,
                            "score": 0,
                            "reasons": ["dom_missing", "resume_inference_failed"]
                        }

                loc_debug = loc_debug or {
                    "method": "none",
                    "confidence": 0.0,
                    "score": 0,
                    "reasons": ["none"]
                }

                if LOCATION_DEBUG:
                    if loc_debug:
                        method = loc_debug.get("method", "unknown")
                        emoji = "🤖" if method == "ner" else ("📄" if method == "dom" else "📐")
                        conf_val = float(loc_debug.get("confidence", 0) or 0)
                        print(f"   {emoji} {method.upper()} score={loc_debug.get('score')} conf={conf_val:.2f}")
                    else:
                        print(f"   🧭 loc(dom)='{location}'" if location else "   🧭 loc: abstain")

                # Screening Q/A
                await click_screening_tab(page)
                await page.wait_for_timeout(SCREENING_WAIT_AFTER_CLICK_MS)
                qa = await get_screening_pairs(page)

                qa_dict = OrderedDict()
                for q_text, a_text in qa:
                    key = canonical_q(q_text)
                    if key not in qa_dict or (not qa_dict[key][1] and a_text):
                        qa_dict[key] = (q_text, a_text)

                screening_qa_text = " ".join([f"{pretty}: {ans}" for _, (pretty, ans) in qa_dict.items() if ans])
                # ─── Combined profile text blob (for semantic tagging / unified keyword scans)
                # Note: pdf_text may already include OCR fallback or OCR_DEBUG_ALWAYS_FIRST_PAGE output
                dom_text = " ".join([name, title, company, location or "", dom_resume_text]).strip()
                screening_text = screening_qa_text
                resume_text = pdf_text
                profile_blob = "\n".join([
                    dom_text or "",
                    screening_text or "",
                    resume_text or "",
                ])
                # Use profile_blob for keyword search (includes all text incl. OCR)
                keyword_hits = find_keyword_hits([profile_blob], results["keywords"])
                kws_cell = ", ".join(keyword_hits)
                # Semantic tags from profile_blob
                semantic_tags_cell = ""
                if SEMANTIC_TAGGING and SEM_TAGGER:
                    try:
                        tags = SEM_TAGGER.predict(profile_blob, threshold=SEMANTIC_THRESHOLD, top_k=SEMANTIC_TOP_K)
                        semantic_tags_cell = ", ".join(f"{t} ({s:.2f})" for t, s in tags)
                    except Exception:
                        pass

                row = {
                    "name": name, "title": title, "company": company,
                    "location": location or "", "url": url, "kws": kws_cell,
                    "semantic_tags": semantic_tags_cell,
                    "_metadata": loc_debug or {}
                }
                
                # Debug logging
                if LOCATION_DEBUG and results["debug_rows"] is not None:
                    results["debug_rows"].append([
                        url, 
                        name, 
                        location or "",
                        (loc_debug or {}).get("method", "unknown"),
                        (loc_debug or {}).get("confidence", ""),
                        (loc_debug or {}).get("score", ""),
                        "+".join((loc_debug or {}).get("reasons", []))
                    ])
                
                for key, (pretty, ans) in qa_dict.items():
                    if key not in results["qcat"]:
                        results["qcat"][key] = pretty
                    results["qcounts"][key] += 1
                    row[key] = ans

                results["rows"].append(row)
                print(f"   ✅ W{wid} Q/A={len(qa_dict)}; loc='{location or ''}'; kws='{kws_cell}'")
                break

            except (PlaywrightTimeout, Exception) as e:
                msg = str(e).splitlines()[0]
                transient = any(t in msg for t in ("Timeout", "closed", "Connection"))
                print(f"   ↻ W{wid} attempt {attempt} error: {msg}")
                if attempt < RETRY_LIMIT and transient:
                    await asyncio.sleep(random.uniform(0.8, 1.5))
                    await dismiss_overlays(page)
                    continue
                results["errs"].append((url, msg))
                break

        processed += 1
        if processed % SWEEP_EVERY == 0:
            # We leave stray PDF tabs alone here; closing aggressively can race with workers
            pass
        q.task_done()

# ─── Main ───────────────────────────────────────────────────────────────────
async def run(
    email_to: Optional[str] = None,
    email_only: bool = False,
    keywords: Optional[List[str]] = None,
):
    global GLOBAL_PDF_EXECUTOR
    keywords = keywords or canonicalize_keywords(DEFAULT_KEYWORDS)
    GLOBAL_PDF_EXECUTOR = concurrent.futures.ThreadPoolExecutor(max_workers=WORKERS * 2)

    async with async_playwright() as p:
        browser = await p.chromium.connect_over_cdp(CDP_URL)
        ctx = browser.contexts[0]
        await setup_context_perf_tuning(ctx)

        def _on_new_page(pg):
            async def _check_and_maybe_close():
                try:
                    await pg.wait_for_load_state("domcontentloaded", timeout=5000)
                    u = (pg.url or "").lower()
                    if any(h in u for h in PDF_URL_HINTS):
                        await asyncio.sleep(0.2)
                        await pg.close()
                except Exception:
                    pass
            asyncio.create_task(_check_and_maybe_close())

        ctx.on("page", _on_new_page)

        # Find SmartRecruiters tab (Chrome may put omnibox popup or other internal tabs first)
        listing = None
        for pg in ctx.pages:
            u = (pg.url or "").lower()
            if "smartrecruiters.com" in u and not u.startswith("chrome://"):
                listing = pg
                break
        if not listing:
            print("⚠️ No SmartRecruiters tab found. Open the prospect list in a tab and run again.")
            await browser.close()
            GLOBAL_PDF_EXECUTOR.shutdown(wait=False, cancel_futures=True)
            return
        await ensure_page_viewport(listing)
        listing.set_default_navigation_timeout(PAGE_LOAD_TIMEOUT)
        await dismiss_overlays(listing)

        role_name, role_number = await get_job_details(listing)
        if role_name or role_number:
            safe_name = _sanitize_filename(role_name) or "Report"
            safe_num = _sanitize_filename(role_number) or ""
            output_filename = f"{safe_name}_{safe_num}.xlsx".strip("_") if safe_num else f"{safe_name}.xlsx"
        else:
            output_filename = OUTFILE
        if role_name or role_number:
            print(f"📋 Job: {role_name or '(unknown)'} | Req: {role_number or '(unknown)'} → {output_filename}")

        kw_preview = ", ".join(keywords[:10])
        if len(keywords) > 10:
            kw_preview += f" ... (+{len(keywords) - 10} more)"
        print(f"🔍 Keywords ({len(keywords)}): {kw_preview}")
        print("⏳ Scrolling to load all profiles...")
        try:
            links = await collect_all_profile_links(listing)
        except Exception as e:
            print(f"⚠️ Error collecting links: {e}")
            await browser.close()
            GLOBAL_PDF_EXECUTOR.shutdown(wait=False, cancel_futures=True)
            return

        if not links:
            print("⚠️ No applicant links found.")
            await browser.close()
            GLOBAL_PDF_EXECUTOR.shutdown(wait=False, cancel_futures=True)
            return

        print(f"📊 Found {len(links)} profiles. Workers: {WORKERS}, Screening wait: {SCREENING_WAIT_AFTER_CLICK_MS}ms")

        results = {
            "rows": [],
            "errs": [],
            "qcat": OrderedDict(),
            "qcounts": defaultdict(int),
            "debug_rows": [] if LOCATION_DEBUG else None,
            "keywords": keywords,
        }

        q: asyncio.Queue = asyncio.Queue()
        for idx, path in enumerate(links, start=1):
            q.put_nowait((idx, len(links), path))
        for _ in range(WORKERS):
            q.put_nowait(None)

        pacer = NavPacer(NAV_MIN_GAP_MS)
        pages = await asyncio.gather(*[ctx.new_page() for _ in range(WORKERS)])
        await asyncio.gather(*[ensure_page_viewport(pg) for pg in pages])
        for pg in pages:
            pg.set_default_timeout(10000)
            pg.on("dialog", lambda d: asyncio.create_task(d.dismiss()))

        workers = [
            asyncio.create_task(worker_loop(i + 1, ctx, pages[i], q, pacer, results))
            for i in range(WORKERS)
        ]

        await q.join()
        for w in workers:
            try:
                await w
            except Exception:
                pass

        for pg in pages:
            try:
                await pg.close()
            except Exception:
                pass
        try:
            await close_stray_pdf_tabs(ctx)
        except Exception:
            pass

        # ─── Excel Output ───────────────────────────────────────────────────
        print("\n📝 Creating Excel report...")
        qkeys = list(results["qcat"].keys())
        qkeys.sort(key=lambda k: (-results["qcounts"][k], results["qcat"][k].lower()))

        wb = Workbook()
        ws = wb.active
        ws.title = "Matches"
        bold = Font(bold=True)

        headers = ["Name", "Job title", "Company", "Location", "ATS Profile link", "Keywords match", "Semantic tags"]
        headers.extend([results["qcat"][k] for k in qkeys])
        ws.append(headers)
        for c in ws[1]:
            c.font = bold

        for r in results["rows"]:
            vals = [r.get("name", ""), r.get("title", ""), r.get("company", ""),
                    r.get("location", ""), "ATS Profile link", r.get("kws", ""), r.get("semantic_tags", "")]
            vals.extend([r.get(k, "") for k in qkeys])
            ws.append(vals)
            rr = ws.max_row
            
            # ATS link
            link_cell = ws.cell(row=rr, column=5)
            link_cell.hyperlink = r["url"]
            link_cell.font = Font(color="0000FF", underline="single")
            
            # Location cell coloring
            loc_cell = ws.cell(row=rr, column=4)
            meta = r.get("_metadata", {})
            if meta.get("method") == "ner":
                # GREEN for NER
                loc_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                loc_cell.font = Font(italic=True, bold=True)
                if loc_cell.value:
                    loc_cell.value = f"🤖 {loc_cell.value}"
            elif meta.get("method") == "regex":
                # BLUE for Regex
                loc_cell.fill = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")
                if loc_cell.value:
                    loc_cell.value = f"📐 {loc_cell.value}"
            elif meta.get("method") == "dom":
                # YELLOW for DOM data (from website)
                loc_cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
                if loc_cell.value:
                    loc_cell.value = f"📄 {loc_cell.value}"

        if results["errs"]:
            wse = wb.create_sheet("Errors")
            wse.append(["ATS Profile link", "ErrorMessage"])
            for c in wse[1]:
                c.font = bold
            for u, m in results["errs"]:
                wse.append([u, m])
                rr = wse.max_row
                lc = wse.cell(row=rr, column=1)
                if u != "Unknown":
                    lc.hyperlink = u
                lc.font = Font(color="0000FF", underline="single")

        # LocationDebug sheet
        if LOCATION_DEBUG and results["debug_rows"]:
            wsd = wb.create_sheet("LocationDebug")
            wsd.append(["ATS Profile link", "Name", "Location", "Method", "Confidence", "Score", "Reasons"])
            for c in wsd[1]:
                c.font = bold
            for rowd in results["debug_rows"]:
                wsd.append(rowd)
                rr = wsd.max_row
                lc = wsd.cell(row=rr, column=1)
                lc.hyperlink = rowd[0]
                lc.font = Font(color="0000FF", underline="single")
                
                # Color code the Method column
                method_cell = wsd.cell(row=rr, column=4)
                method_val = str(rowd[3]).lower() if len(rowd) > 3 else ""
                if "ner" in method_val:
                    method_cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                elif "regex" in method_val:
                    method_cell.fill = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")
                elif "dom" in method_val:
                    method_cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

        ws.freeze_panes = "A2"
        last_col = get_column_letter(ws.max_column)
        last_row = ws.max_row
        if last_row >= 2:
            tbl = Table(displayName="MatchesTable", ref=f"A1:{last_col}{last_row}")
            tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
            ws.add_table(tbl)

        widths = {"A": 28, "B": 28, "C": 34, "D": 26, "E": 26, "F": 22, "G": 36}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        wrap = Alignment(wrap_text=True, vertical="top")
        for col_idx in range(7, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 36
            for rr in range(2, ws.max_row + 1):
                ws.cell(row=rr, column=col_idx).alignment = wrap

        print(f"\n📊 Summary:")
        print(f"  ✅ Processed: {len(results['rows'])} profiles")
        print(f"  ❌ Errors: {len(results['errs'])}")
        print(f"  ❓ Unique questions: {len(results['qcat'])}")

        # Save to buffer for email; optionally to disk
        buffer = io.BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()

        if email_to:
            if role_name and role_number:
                email_subject = f"SmartRecruiters Prospect Report - {role_name} ({role_number})"
            elif role_name:
                email_subject = f"SmartRecruiters Prospect Report - {role_name}"
            elif role_number:
                email_subject = f"SmartRecruiters Prospect Report - {role_number}"
            else:
                email_subject = "SmartRecruiters Prospect Report"
            sent = send_email(email_to, excel_bytes, filename=output_filename, subject=email_subject)
            if email_only:
                if sent:
                    print(f"\n🎯 Done! Report emailed to {email_to} (not saved locally)")
                else:
                    wb.save(output_filename)
                    print(f"\n⚠️  Email failed. Report saved to: {output_filename}")
            else:
                wb.save(output_filename)
                print(f"\n🎯 Done! Results saved to: {output_filename} and emailed to {email_to}")
        else:
            wb.save(output_filename)
            print(f"\n🎯 Done! Results saved to: {output_filename}")

        await browser.close()
        GLOBAL_PDF_EXECUTOR.shutdown(wait=False, cancel_futures=True)

if __name__ == "__main__":
    _load_email_credentials()
    parser = argparse.ArgumentParser(description="SmartRecruiters prospect scraper")
    parser.add_argument(
        "--email-to",
        type=str,
        metavar="EMAIL",
        help="Email report to this address (default: EMAIL_TO or EMAIL_USER from env)",
    )
    parser.add_argument(
        "--save-local",
        action="store_true",
        help="Also save report locally (default: email only, no local file)",
    )
    parser.add_argument(
        "--keywords",
        type=str,
        metavar="LIST",
        help="Keywords to search (Ctrl+F style). Comma-separated, e.g. 'Python, ML, AWS, Chennai'",
    )
    parser.add_argument(
        "--keywords-file",
        type=str,
        metavar="PATH",
        help="Load keywords from file (one per line or comma-separated)",
    )
    args = parser.parse_args()
    email_to = args.email_to or os.environ.get("EMAIL_TO") or os.environ.get("EMAIL_USER")
    email_only = not args.save_local
    if email_only and not email_to:
        parser.error("Email-only mode requires --email-to or EMAIL_TO/EMAIL_USER in env")

    keywords = resolve_keywords(args.keywords, args.keywords_file)

    async def _run():
        await run(email_to=email_to or None, email_only=email_only, keywords=keywords)

    asyncio.run(_run())